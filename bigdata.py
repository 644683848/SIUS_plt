import csv

import requests
import json
import xlsxwriter
from datetime import datetime, timedelta
from openpyxl import load_workbook

# URL endpoint
url = 'http://admin.hbshejishejian.com/api/train/bullet_map_detail'

# Headers
headers = {
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-US;q=0.7,ru;q=0.6',
    'Connection': 'keep-alive',
    'Content-Type': 'application/json;charset=UTF-8',
    'Origin': 'http://db.hbshejishejian.com',
    'Referer': 'http://db.hbshejishejian.com/',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
    'xx-api-timestamp': '1712911825',
    'xx-api-token': '16be6f7c3945c34be8a1952a0f67d5da',
    'xx-api-user': '22'
}

day = "2024-04-05"
posture_map = {
    '0': '跪射',
    '1': '卧射',
    '2': '立射'
}

# Request body
payload = {
    "coach_id": "11",
    "project2": "26",
    "day": day
}


def generate_date_list(start_date_str, num_days):
    date_format = "%Y-%m-%d"
    start_date = datetime.strptime(start_date_str, date_format)
    date_list = [start_date.strftime(date_format)]
    for i in range(1, num_days):
        next_date = start_date + timedelta(days=i)
        date_list.append(next_date.strftime(date_format))
    return date_list


def append_to_csv(file_path, data):
    with open(file_path, 'a', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(data)


if __name__ == '__main__':
    # Convert payload to JSON string
    json_payload = json.dumps(payload)

    # Sending POST request
    response = requests.post(url, headers=headers, data=json_payload)

    result = {}
    data = response.json()['data']
    for posture, items in data.items():
        if posture not in ['0', '1', '2']:
            break
        name = items['member_name']
        train_infos = items['train_info']
        for train_info in train_infos:
            key = name + ';' + day + ';' + train_info['shoot_type'] + ';' + posture_map[posture]
            if result.get(key) is None:
                result[key] = []
            result[key].append(str(train_info['avg_circle']))

    file_path = r'C:\Users\64468\PycharmProjects\SIUS\example.xlsx'

    rows = []
    for key, val in result.items():
        row = key.split(';')
        row.extend(val)
        rows.append(row)

    wb = load_workbook(file_path)
    # Select First Worksheet
    ws = wb.worksheets[0]

    # Append 2 new Rows - Columns A - D
    for row_data in rows:
        # Append Row Values
        ws.append(row_data)

    wb.save(file_path)

    # with xlsxwriter.Workbook(file_path) as workbook:
    #     worksheet = workbook.add_worksheet('Sheet1')
    #     # Get the current number of rows in the worksheet
    #     existing_rows_count = worksheet.dim_rowmax + 1 if worksheet.dim_rowmax else 0
    #     # Append new data
    #     for row_num, data in enumerate(rows):
    #         # Calculate the row number to write by adding the existing row count
    #         worksheet.write_row(existing_rows_count + row_num, 0, data)
